#!/usr/bin/env python
# coding: utf-8



import numpy as np
import pandas as pd
from openpyxl import load_workbook

#  load data, and seperate GMV, Sales, and ATV


class DFData():
    """
    读取和保存数据
    """
    def __init__(self, path, unit=1, index=None, groupby=None):

        data = pd.read_excel(path, index_col=index)
        data['gmv'] = data['gmv'] / unit
        data['sale_qtty'] = data['sale_qtty'] / unit

        self.data = data.groupby(groupby).sum()
        self.gmv = self.data['gmv'].unstack()
        self.sales = self.data['sale_qtty'].unstack()
        self.atv = self.gmv / self.sales

        if 'source' in data.columns and "cat" in path:
            groupby.insert(1,'source')
            self.cid_data = data.groupby(groupby).sum()
            self.pop_gmv = self.cid_data.loc[(slice(None),'pop'),'gmv'].unstack().reset_index(level='source',drop=True)
            self.self_gmv = self.cid_data.loc[(slice(None), 'self'), 'gmv'].unstack().reset_index(level='source',drop=True)
        else:
            self.cid_data = None
            self.pop_gmv = None
            self.self_gmv = None

# load brands or catalog list


def load_cat(path, name):
    """
    导入品牌目录，并根据品类（name）获取品牌列表
    """
    brands = pd.read_excel(path).fillna(0)
    brand_lit = brands[name].tolist()
    while 0 in brand_lit:
        brand_lit.remove(0)
    return brand_lit


# calculate the growth


def yoy_growth(df, period):
    """计算yoy增长率
    """
    tmp = df.to_period(period).groupby('dt').sum()
    tmp1 = tmp['2018']
    tmp2 = tmp['2019']
    growth = tmp2.copy()
    m, n = tmp2.shape
    for i in range(n):
        for j in range(m):
            growth.iloc[j, i] = (tmp2.iloc[j, i] / tmp1.iloc[j, i] - 1)
    return growth


def cr_growth(df, period):
    """
    计算环比增长率
    """
    tmp = df.to_period(period).groupby('dt').sum()
    growth = tmp.copy()
    m, n = tmp.shape
    for i in range(n):
        growth.iloc[:, i] = (tmp.iloc[:, i] / tmp.iloc[:, i].shift(1) - 1)
    return growth


#  save datas


def bubble_data(gmv, date, kind=None, cat_gmv=None, brands=None):
    """
        格式化需要画泡泡图的数据
    input:  gmv（df）：需要画泡泡图的数据
            date(str): 日期 ‘2019-04-01’
            kind(str): 数据的类型（cat = '分品类'，其他表示具体的品牌数据）
            cat_gmv(df): 品类的总数据；当cat为品牌时，必选。
            brands(str)：品类的名称，当cat为品牌时，必选。
    output: bubble(df): 返回泡泡数据，份额、增速和gmv
    """

    if kind == "大家电" or kind == "小家电":
        kind = '家用电器'
    if kind == "cat":
        data = gmv.loc[date, :]
        total = data.sum()
        gmv_pct = data / total
        gmv_yoy = yoy_growth(gmv, "M").loc[date, :]
    else:
        total = cat_gmv.loc[date, kind]
        data = gmv[brands].loc[date, :]
        gmv_pct = data / total
        gmv_yoy = yoy_growth(gmv, "M")[brands].loc[date, :]
    bubble = pd.DataFrame(gmv_pct)
    bubble.columns = ['平均市场份额']
    bubble["销售额增速"] = np.array(gmv_yoy)
    bubble["销售额"] = np.array(data)
    return bubble


def save_bubble(gmv, cat, catalog, catlist, date, bubblepath, file_path):
    """
    保存bubble数据
    """
    writer = pd.ExcelWriter(file_path + date + bubblepath)
    if cat.cid_data is None:
        cat_bubble = bubble_data(cat.gmv, date, kind='cat')
        cat_bubble.to_excel(writer, sheet_name='平台')
    if cat.cid_data is not None:
        cat_pop = bubble_data(cat.pop_gmv, date, kind='cat')
        cat_self = bubble_data(cat.self_gmv, date, kind='cat')
        cat_pop.to_excel(writer, sheet_name='平台')
        cat_self.to_excel(writer, sheet_name='自营')
    for kind in catlist:
        brands = load_cat(catalog, kind)
        b_brand = bubble_data(
            gmv,
            date,
            kind=kind,
            cat_gmv=cat.gmv,
            brands=brands)
        b_brand.to_excel(writer, sheet_name=kind)
    writer.save()
    writer.close()


def save_cat_sheet(df, path, name):
    '''保存品类数据
    input: df（df）: 需要整理的原始数据
            path(str)：保存的路径
            name: sheet 名
    '''
    n = df.shape[0]
    my_growth = yoy_growth(df, 'M')
    qy_growth = yoy_growth(df, "Q")
    c_growth = cr_growth(df, "M")

    workbook = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = workbook
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
        workbook.save(path)
    df.to_excel(writer, sheet_name=name)
    my_growth.to_excel(writer, sheet_name=name, startrow=n + 3)
    qy_growth.to_excel(writer, sheet_name=name, startrow=2 * n + 3)
    c_growth.to_excel(writer, sheet_name=name, startrow=3 * n + 3)
    writer.save()
    writer.close()





def save_brand(cat, path, cat_path, gmv, sales, atv, gmv_g, sales_g, atv_g):
    '''保存品牌数据
    input: cat（str）: 需要整理的品类的名称
            path(str)：保存的路径
    '''
    brands = load_cat(cat_path, cat)
    gmv = gmv[brands]
    sales = sales[brands]
    atv = atv[brands]
    gmv_g = gmv_g[brands]
    sales_g = sales_g[brands]
    atv_g = atv_g[brands]

    m, n = gmv.shape
    workbook = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = workbook
    if cat in workbook.sheetnames:
        workbook.remove(workbook[cat])
        workbook.save(path)

    gmv.to_excel(writer, sheet_name=cat)
    gmv_g.to_excel(writer, sheet_name=cat, startcol=n + 3)

    sales.to_excel(writer, sheet_name=cat, startrow=m + 3)
    sales_g.to_excel(writer, sheet_name=cat, startrow=m + 3, startcol=n + 3)

    atv.to_excel(writer, sheet_name=cat, startrow=2 * m + 6)
    atv_g.to_excel(writer, sheet_name=cat, startrow=2 * m + 6, startcol=n + 3)
    writer.save()
    writer.close()

def format_data(cat,brand,brand_path, cat_path,cat_name_path,bubblepath,catlist,plat,date,unit):

    print("----导入数据----")

    # 2. 计算增长率
    gmv_g = yoy_growth(brand.gmv, "M")
    sales_g= yoy_growth(brand.sales, "M")
    atv_g = yoy_growth(brand.atv, "M")
    print("----增长率----")

    # 3. 保存品类数据
    save_cat_sheet(cat.gmv, cat_path, 'gmv')
    save_cat_sheet(cat.sales, cat_path, 'sales')
    save_cat_sheet(cat.atv, cat_path, 'atv')

    if cat.cid_data is not None:
        save_cat_sheet(cat.pop_gmv, cat_path, '平台')
        save_cat_sheet(cat.self_gmv, cat_path, '自营')

    print("----品类----")

    # 4. 保存品牌数据
    for cat_name in catlist:
        save_brand(
            cat_name,
            brand_path,
            cat_name_path,
            brand.gmv,
            brand.sales,
            brand.atv,
            gmv_g,
            sales_g,
            atv_g)
    print("----品牌----")
    # 5. 保存泡泡图格式数据
    save_bubble(
        brand.gmv,
        cat,
        cat_name_path,
        catlist,
        date,
        bubblepath,
        plat)
    print("----完成----")

# * 运行
if __name__ == '__main__':

    unit = 100000000
    date = '2019-05-01'
    #Tmall
    print("----------Tmall------------")
    brand_path_T = "./Tmall/Tmall_brand.xlsx"
    cat_path_T = './Tmall/Tmall_cat.xlsx'
    catlist_T = ['医药保健', '酒类', '大家电', '小家电', '美妆个护', '服装鞋包']
    cat_name_path_T = './Tmall/brand_catalog.xlsx'
    bubblepath_T = 'bubble_data_Tmall.xlsx'
    plat_T = "./Tmall/"

    T_cat = DFData(cat_path_T, index='dt', groupby=['dt', 'cid1_name'], unit=unit)
    T_brand = DFData(brand_path_T, index='dt', groupby=['dt', 'main_brand_name'], unit=unit)
    format_data(T_cat,T_brand,brand_path_T, cat_path_T, cat_name_path_T, bubblepath_T, catlist_T, plat_T, date, unit)

    #JD
    print("----------JD------------")
    catlist_J = ['医药保健', '酒类', '大家电', '小家电', '食品饮料及生鲜']
    brand_path_J = "./JD/JD_brands.xlsx"
    cat_path_J = './JD/JD_cats.xlsx'
    cat_name_path_J = './JD/JD_catalog.xlsx'
    bubblepath_J = 'bubble_data_JD.xlsx'
    plat_J = "./JD/"

    J_cat = DFData(cat_path_J, index='dt', groupby=['dt', 'cid1_name'], unit=unit)
    J_brand = DFData(brand_path_J, index='dt', groupby=['dt', 'main_brand_name'], unit=unit)
    format_data(J_cat,J_brand,brand_path_J, cat_path_J, cat_name_path_J, bubblepath_J, catlist_J, plat_J,date,unit)

    # 加总
    catlist_total = ['酒类', '大家电', '小家电']
    gmv =
    for cat_name in catlist_total:
        save_brand(
            cat_name,
            './total_brand.xlsx',
            cat_name_path_T,
            brand.gmv,
            brand.sales,
            brand.atv,
            gmv_g,
            sales_g,
            atv_g)